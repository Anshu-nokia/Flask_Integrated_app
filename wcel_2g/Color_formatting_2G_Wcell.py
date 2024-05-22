from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import PatternFill
import pandas as pd
import time
from openpyxl.styles import Font
import sys
import warnings
from datetime import datetime, timedelta


# Ignore all warnings
warnings.filterwarnings("ignore")



# Get the current date and time
now = datetime.now()
last_day = datetime.now().date() - timedelta(days=1)
last_day_date = last_day.strftime('%d-%b-%y')
df = pd.read_excel("/home/CNO_Server/scripts/Wcell_2G_Tracker/ZM_2G_WCL.xlsm", index_col=False, sheet_name="CSSR", skiprows=4)
column_name = df.columns[36]
column_name = pd.to_datetime(column_name)
column_name = column_name.strftime('%d-%b-%y')
# print(column_name, last_day_date)
if column_name != last_day_date:
    print("Tracker aren't updated. Exiting...")
    sys.exit()  # This will terminate the program





start_time = time.time()

existing_file_path = r"/home/CNO_Server/scripts/Wcell_2G_Tracker/ZM_2G_WCL.xlsm"
wb = load_workbook(existing_file_path, read_only=False, keep_vba=True, keep_links=True)

column_list = ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']
sheet_list = ['CSSR', 'TCH_Drop', 'SDCCH_Drop', 'SDCCH_Cong', 'TCH_Cong', 'RNA_BBH', 'RNA_24HR']
# Iterate over the dictionary using a for loop
for value in sheet_list:
    sheet_name = wb[value]
    print(f"Updating {value}...")
    # source_columns = range(8, 38)  # Assuming (30 columns)
    # target_columns = range(7, 37)  # Assuming (30 columns)
    # Find the last row in column J
    df1 = pd.read_excel(existing_file_path, skiprows=4,
                        sheet_name=value)
    # Define the range of cells in column AK you want to check
    start_row = 6
    end_row = len(df1)  # Adjust the end row as needed
    print("Max_row:", end_row)

    for row in range(start_row, end_row + 1):
        for i in column_list:
            # print(f"Updating {i} Column...")
            cell = sheet_name[f'{i}{row}']  # Get the cell in column AK for the current row
            status_cell = sheet_name[f'F{row}']
            # print(status_cell.value)
            threshold = None
            # condition = False  # Initialize condition variable
    
            ##---taking threshold acording to sheet name for conditional formating -------------
    
            if value == "CSSR":
                if status_cell.value == "Platinum" or status_cell.value == "Gold":
                    threshold = 98.5
                elif status_cell.value == "Silver":
                    threshold = 96
            elif value == "TCH_Drop":
                if status_cell.value == "Platinum" or status_cell.value == "Gold":
                    threshold = 1.75
                elif status_cell.value == "Silver":
                    threshold = 2
            elif value == "SDCCH_Drop":
                if status_cell.value == "Platinum" or status_cell.value == "Gold":
                    threshold = 1.5
                elif status_cell.value == "Silver":
                    threshold = 1.75
            elif value == "SDCCH_Cong":
                if status_cell.value == "Platinum" or status_cell.value == "Gold" or status_cell.value == "Silver":
                    threshold = 0.25
            elif value == "TCH_Cong":
                if status_cell.value == "Platinum" or status_cell.value == "Gold" or status_cell.value == "Silver":
                    threshold = 1
            elif value == "RNA_BBH" or value == "RNA_24HR":
                if status_cell.value == "Platinum" or status_cell.value == "Gold":
                    threshold = 98.5
                elif status_cell.value == "Silver":
                    threshold = 97.2
    
            try:
                ##----comparison with >=
                if value == "CSSR" or value == "RNA_BBH" or value == "RNA_24HR":
                    if cell.value is not None and not isinstance(cell.value, str) and threshold is not None and cell.value >= threshold:
                        ##Green
                        # print("Green")
                        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type="solid")
                        cell.font = Font(name='Calibri', size=8, color="000000")  # Set font color to black
                        
                    else:
                        ##Red
                        # print("Red")
                        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type="solid")
                        cell.font = Font(name='Calibri', size=8, color="FFFFFF", bold=True)  # Set font color to white
                ##----comparison with <=
                elif value == "TCH_Drop" or value == "SDCCH_Drop" or value == "SDCCH_Cong" or value == "TCH_Cong":
                    if cell.value is not None and not isinstance(cell.value, str) and threshold is not None and cell.value <= threshold:
                        ##Green
                        # print("Green")
                        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type="solid")
                        cell.font = Font(name='Calibri', size=8, color="000000")  # Set font color to black
                    else:
                        ##Red
                        # print("Red")
                        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type="solid")
                        cell.font = Font(name='Calibri', size=8, color="FFFFFF", bold=True)  # Set font color to white
    
            except ValueError:
                pass



wb.save(r"/home/CNO_Server/scripts/Wcell_2G_Tracker/ZM_2G_WCL.xlsm")

end_time = time.time()
print("Time: ", end_time - start_time)
# Reset warnings to default behavior
warnings.resetwarnings()









